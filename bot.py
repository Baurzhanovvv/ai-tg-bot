import asyncio
import logging
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict

import aiohttp
from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import Message, ReplyKeyboardMarkup, KeyboardButton, FSInputFile
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pydub import AudioSegment

# Загрузка переменных окружения
load_dotenv()

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Конфигурация
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
OPENROUTER_MODEL = os.getenv("OPENROUTER_MODEL", "anthropic/claude-3.5-haiku")
OPENROUTER_API_URL = "https://openrouter.ai/api/v1/chat/completions"
GROQ_API_KEY = os.getenv("GROQ_API_KEY")
GROQ_WHISPER_URL = "https://api.groq.com/openai/v1/audio/transcriptions"
PROMPT_FILE = "prompt.md"
MAX_HISTORY_MESSAGES = int(os.getenv("MAX_HISTORY_MESSAGES", "10"))  # Максимум сообщений в истории

# Глобальная переменная для системного промпта
SYSTEM_PROMPT: Optional[str] = None

# История диалогов для каждого пользователя: {user_id: [{"role": "user/assistant", "content": "..."}]}
user_conversations: Dict[int, List[Dict[str, str]]] = defaultdict(list)


def load_system_prompt() -> bool:
    """Загружает системный промпт из файла prompt.md"""
    global SYSTEM_PROMPT
    
    try:
        prompt_path = Path(PROMPT_FILE)
        if not prompt_path.exists():
            logger.error(f"Файл {PROMPT_FILE} не найден!")
            return False
        
        SYSTEM_PROMPT = prompt_path.read_text(encoding='utf-8').strip()
        logger.info(f"Системный промпт успешно загружен из {PROMPT_FILE}")
        return True
    except Exception as e:
        logger.error(f"Ошибка при загрузке системного промпта: {e}")
        return False


async def transcribe_audio(audio_path: str) -> Optional[str]:
    """
    Транскрибирует аудиофайл в текст используя Groq Whisper API (БЕСПЛАТНО!).
    
    Args:
        audio_path: Путь к аудиофайлу
        
    Returns:
        Транскрибированный текст или None в случае ошибки
    """
    if not GROQ_API_KEY or GROQ_API_KEY == "your_groq_api_key_here":
        logger.error("Groq API ключ не установлен! Голосовые сообщения недоступны.")
        return None
    
    mp3_path = None
    try:
        # Конвертируем аудио в формат MP3 для Whisper
        audio = AudioSegment.from_file(audio_path)
        mp3_path = audio_path.replace('.oga', '.mp3')
        audio.export(mp3_path, format="mp3")
        logger.info(f"Аудио сконвертировано в MP3: {mp3_path}")
        
        # Отправляем запрос к OpenAI Whisper API
        async with aiohttp.ClientSession() as session:
            with open(mp3_path, 'rb') as audio_file:
                form_data = aiohttp.FormData()
                form_data.add_field('file', audio_file, filename='audio.mp3', content_type='audio/mpeg')
                form_data.add_field('model', 'whisper-large-v3')
                form_data.add_field('language', 'ru')  # Русский язык (можно удалить для автоопределения)
                
                headers = {
                    'Authorization': f'Bearer {GROQ_API_KEY}'
                }
                
                async with session.post(
                    GROQ_WHISPER_URL,
                    headers=headers,
                    data=form_data
                ) as response:
                    if response.status == 200:
                        result = await response.json()
                        text = result.get('text', '').strip()
                        logger.info(f"Аудио успешно транскрибировано через Groq: {text[:100]}...")
                        
                        # Удаляем временные файлы
                        os.remove(audio_path)
                        os.remove(mp3_path)
                        
                        return text
                    else:
                        error_text = await response.text()
                        logger.error(f"Ошибка транскрипции Groq Whisper API: {response.status} - {error_text}")
                        return None
                        
    except Exception as e:
        logger.error(f"Ошибка при транскрибации аудио: {e}")
        # Очищаем временные файлы в случае ошибки
        try:
            if audio_path and os.path.exists(audio_path):
                os.remove(audio_path)
            if mp3_path and os.path.exists(mp3_path):
                os.remove(mp3_path)
        except Exception as cleanup_error:
            logger.error(f"Ошибка при удалении временных файлов: {cleanup_error}")
        return None


def add_to_history(user_id: int, role: str, content: str) -> None:
    """
    Добавляет сообщение в историю пользователя.
    
    Args:
        user_id: ID пользователя Telegram
        role: Роль ('user' или 'assistant')
        content: Текст сообщения
    """
    user_conversations[user_id].append({"role": role, "content": content})
    
    # Ограничиваем историю последними N сообщениями (не считая системный промпт)
    if len(user_conversations[user_id]) > MAX_HISTORY_MESSAGES:
        user_conversations[user_id] = user_conversations[user_id][-MAX_HISTORY_MESSAGES:]
    
    logger.info(f"История пользователя {user_id}: {len(user_conversations[user_id])} сообщений")


def clear_history(user_id: int) -> None:
    """
    Очищает историю диалога пользователя.
    
    Args:
        user_id: ID пользователя Telegram
    """
    user_conversations[user_id] = []
    logger.info(f"История пользователя {user_id} очищена")


def get_conversation_history(user_id: int) -> List[Dict[str, str]]:
    """
    Получает историю диалога пользователя.
    
    Args:
        user_id: ID пользователя Telegram
        
    Returns:
        Список сообщений в формате [{"role": "user/assistant", "content": "..."}]
    """
    return user_conversations[user_id]


def create_main_keyboard() -> ReplyKeyboardMarkup:
    """
    Создает основную клавиатуру с кнопками управления.
    
    Returns:
        ReplyKeyboardMarkup с кнопками
    """
    keyboard = ReplyKeyboardMarkup(
        keyboard=[
            [
                KeyboardButton(text="➡️ Следующий ученик"),
                KeyboardButton(text="📊 Экспорт в Excel")
            ]
        ],
        resize_keyboard=True,
        one_time_keyboard=False
    )
    return keyboard


async def export_to_excel(user_id: int) -> Optional[str]:
    """
    Экспортирует финальный отчёт (8 пунктов) в Excel файл.
    
    Args:
        user_id: ID пользователя Telegram
        
    Returns:
        Путь к созданному файлу или None в случае ошибки
    """
    history = get_conversation_history(user_id)
    
    if not history:
        return None
    
    try:
        # Ищем последний ответ бота с финальным отчётом
        final_report = None
        student_name = None
        
        # Ищем имя ученика в истории
        for msg in history:
            if msg['role'] == 'user':
                # Ищем упоминание имени после вопросов об имени
                content_lower = msg['content'].lower()
                if any(word in content_lower for word in ['зовут', 'имя', 'ученик', 'ученица']):
                    # Пытаемся извлечь имя (первое слово с заглавной буквы после ключевых слов)
                    words = msg['content'].split()
                    for i, word in enumerate(words):
                        if word and word[0].isupper() and len(word) > 2 and word.isalpha():
                            student_name = word
                            break
        
        for msg in reversed(history):
            if msg['role'] == 'assistant':
                # Проверяем что это отчёт с 8 пунктами
                if '1.' in msg['content'] and '8.' in msg['content']:
                    final_report = msg['content']
                    break
        
        if not final_report:
            return None
        
        # Создаем новую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет преподавателя"
        
        # Настройка стилей
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_font = Font(size=11)
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок документа
        ws.merge_cells('A1:B1')
        title = f"Отчет преподавателя"
        if student_name:
            title += f" - {student_name}"
        title += f" - {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        
        # Заголовки таблицы
        ws['A2'] = "Пункт отчёта"
        ws['B2'] = "Комментарий"
        
        for cell in ['A2', 'B2']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
            ws[cell].alignment = header_alignment
            ws[cell].border = border
        
        ws.row_dimensions[2].height = 30
        
        # Ширина столбцов
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 80
        
        # Парсим отчёт на 8 пунктов
        row = 3
        
        # Убираем markdown форматирование (** и т.д.)
        clean_report = re.sub(r'\*\*', '', final_report)
        
        # Ищем все пункты с номерами от 1 до 8
        logger.info(f"Начало парсинга отчёта для пользователя {user_id}")
        logger.info(f"Длина отчёта: {len(clean_report)} символов")
        
        # Разбиваем по паттерну "число."
        points = re.split(r'\n(?=\d+\.)', clean_report)
        logger.info(f"Найдено частей после split: {len(points)}")
        
        for i, point in enumerate(points, 1):
            point = point.strip()
            if not point:
                logger.info(f"Пункт {i} пустой, пропускаем")
                continue
            
            logger.info(f"Обработка пункта {i}: {point[:100]}...")
            
            # Извлекаем заголовок пункта и содержимое
            # Паттерн: "1. Заголовок" или "1. Заголовок:" далее содержимое
            match = re.match(r'^(\d+)\.\s*([^:\n]+):?\s*(.*)', point, re.DOTALL)
            if match:
                point_num = match.group(1)
                point_title = match.group(2).strip()
                point_content = match.group(3).strip()
                
                logger.info(f"Найден пункт #{point_num}: {point_title}")
                
                # Заголовок пункта
                ws[f'A{row}'] = f"{point_num}. {point_title}"
                ws[f'A{row}'].font = Font(bold=True, size=11)
                ws[f'A{row}'].alignment = cell_alignment
                ws[f'A{row}'].border = border
                
                # Содержимое пункта
                ws[f'B{row}'] = point_content
                ws[f'B{row}'].font = cell_font
                ws[f'B{row}'].alignment = cell_alignment
                ws[f'B{row}'].border = border
                
                # Автоматическая высота строки
                ws.row_dimensions[row].height = max(60, len(point_content) // 4)
                
                row += 1
            else:
                logger.warning(f"Пункт {i} не совпал с паттерном: {point[:100]}")
        
        logger.info(f"Всего добавлено строк в Excel: {row - 3}")
        
        # Сохраняем файл с именем формата: ОТЧЕТ_ФАМИЛИЯ_ИМЯ.xlsx
        if student_name:
            # Преобразуем имя в заглавные буквы для имени файла
            student_name_upper = student_name.upper()
            filename = f"ОТЧЕТ_{student_name_upper}.xlsx"
        else:
            # Если имя не найдено, используем ID и дату
            filename = f"ОТЧЕТ_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb.save(filename)
        
        logger.info(f"Создан Excel файл: {filename}")
        return filename
        
    except Exception as e:
        logger.error(f"Ошибка при создании Excel файла: {e}")
        return None


async def get_llm_response(user_id: int, user_message: str) -> Optional[str]:
    """
    Отправляет запрос к OpenRouter API и получает ответ от LLM с учетом истории диалога.
    
    Args:
        user_id: ID пользователя Telegram
        user_message: Текст сообщения пользователя
        
    Returns:
        Ответ от LLM или None в случае ошибки
    """
    if not SYSTEM_PROMPT:
        logger.error("Системный промпт не загружен!")
        return None
    
    try:
        # Добавляем новое сообщение пользователя в историю
        add_to_history(user_id, "user", user_message)
        
        # Получаем историю диалога
        conversation_history = get_conversation_history(user_id)
        
        headers = {
            'Authorization': f'Bearer {OPENROUTER_API_KEY}',
            'Content-Type': 'application/json',
            'HTTP-Referer': 'https://github.com/your-repo',
            'X-Title': 'Telegram AI Bot'
        }
        
        # Формируем сообщения: системный промпт + история диалога
        messages = [
            {'role': 'system', 'content': SYSTEM_PROMPT}
        ] + conversation_history
        
        payload = {
            'model': OPENROUTER_MODEL,
            'messages': messages,
            'max_tokens': 4000  # Увеличили для длинных ответов
        }
        
        async with aiohttp.ClientSession() as session:
            async with session.post(
                OPENROUTER_API_URL,
                headers=headers,
                json=payload
            ) as response:
                if response.status == 200:
                    data = await response.json()
                    answer = data['choices'][0]['message']['content']
                    logger.info(f"Получен ответ от LLM (длина: {len(answer)} символов)")
                    
                    # Добавляем ответ ассистента в историю
                    add_to_history(user_id, "assistant", answer)
                    
                    return answer
                else:
                    error_text = await response.text()
                    logger.error(f"Ошибка API OpenRouter: {response.status} - {error_text}")
                    return None
                    
    except Exception as e:
        logger.error(f"Ошибка при обращении к LLM: {e}")
        return None


# Инициализация бота и диспетчера
bot = Bot(token=TELEGRAM_BOT_TOKEN)
dp = Dispatcher()


@dp.message(Command("start"))
async def cmd_start(message: Message):
    """Обработчик команды /start"""
    user_id = message.from_user.id
    clear_history(user_id)  # Очищаем историю при старте
    
    # Приветствие
    greeting = (
        "👋 Привет! Я ИИ-ассистент преподавателей образовательного центра «Логос».\n\n"
        "📝 Я помогу вам составить структурированный отчёт для родителей.\n\n"
        "🎤 **Чтобы сделать отчет, запишите голосовым сообщением ваше впечатление о работе ученика по следующим пунктам.**\n"
    )
    
    await message.answer(greeting, parse_mode="Markdown", reply_markup=create_main_keyboard())
    
    # Структура отчёта
    structure_text = (
        "📋 **Структура отчёта (8 пунктов):**\n\n"
        "**1. Работа ученика на занятиях.** Общее впечатление за месяц "
        "(вовлеченность в процесс занятия, каким образом проявлял активность за месяц)\n\n"
        "**2. Работа с домашними заданиями** (впечатление от качества выполнения домашних заданий за месяц)\n\n"
        "**3. Комментарий к экзаменационной работе**\n\n"
        "**4. Ожидаемый результат на этот месяц**\n\n"
        "**5. Причины отсутствия прироста и неудовлетворительного результата**\n\n"
        "**6. Рекомендации на будущий месяц ребёнку**\n\n"
        "**7. Рекомендации родителям**\n\n"
        "**8. Дополнительные комментарии**\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "⚠️ **Важно:**\n"
        "• Обязательно скажите, **про кого идет речь** и **какой месяц**.\n"
        "• Отчёт получится лучше, если будете записывать **с экзаменационной работой на руках**.\n"
        "• Старайтесь рассказывать подробно **в баллах** и **в номерах заданий** — обязательно упомяните."
    )
    
    await message.answer(structure_text, parse_mode="Markdown")


@dp.message(F.text == "➡️ Следующий ученик")
async def handle_clear_history(message: Message):
    """Обработчик кнопки перехода к следующему ученику"""
    user_id = message.from_user.id
    clear_history(user_id)
    
    await message.answer(
        "✅ Переходим к следующему ученику!\n\n"
        "История предыдущего отчёта очищена.",
        reply_markup=create_main_keyboard()
    )
    
    # Показываем структуру отчёта снова
    structure_text = (
        "📋 **Структура отчёта (8 пунктов):**\n\n"
        "1. Работа ученика на занятиях\n"
        "2. Работа с домашними заданиями\n"
        "3. Комментарий к экзаменационной работе\n"
        "4. Ожидаемый результат на этот месяц\n"
        "5. Причины отсутствия прироста\n"
        "6. Рекомендации ребёнку\n"
        "7. Рекомендации родителям\n"
        "8. Дополнительные комментарии\n\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "💬 **Какой месяц отчёта?**\n"
        "💬 **Как зовут ученика?**"
    )
    
    await message.answer(structure_text, parse_mode="Markdown")


@dp.message(F.text == "📊 Экспорт в Excel")
async def handle_export_excel(message: Message):
    """Обработчик кнопки экспорта в Excel"""
    user_id = message.from_user.id
    
    history = get_conversation_history(user_id)
    
    logger.info(f"Экспорт Excel для пользователя {user_id}, история: {len(history)} сообщений")
    
    if not history:
        await message.answer(
            "❌ История пуста! Нечего экспортировать.",
            reply_markup=create_main_keyboard()
        )
        return
    
    # Проверяем наличие финального отчёта
    has_final_report = False
    for msg in reversed(history):
        if msg['role'] == 'assistant' and '1.' in msg['content'] and '8.' in msg['content']:
            has_final_report = True
            logger.info(f"Найден финальный отчёт длиной {len(msg['content'])} символов")
            logger.info(f"Первые 200 символов: {msg['content'][:200]}")
            break
    
    if not has_final_report:
        await message.answer(
            "❌ Финальный отчёт еще не создан!\n\n"
            "Пожалуйста, заполните все 8 пунктов отчёта в диалоге с ботом, "
            "затем попробуйте экспорт снова.",
            reply_markup=create_main_keyboard()
        )
        return
    
    await message.answer("⏳ Формирую отчёт в Excel, подождите...")
    
    # Создаем Excel файл
    filename = await export_to_excel(user_id)
    
    if filename:
        try:
            # Извлекаем имя ученика из названия файла для подписи
            student_name_from_file = None
            if filename.startswith("ОТЧЕТ_") and not filename.split("_")[1].isdigit():
                student_name_from_file = filename.replace("ОТЧЕТ_", "").replace(".xlsx", "")
            
            # Формируем красивую подпись
            caption = "📊 Отчёт преподавателя"
            if student_name_from_file:
                caption += f" - {student_name_from_file}"
            caption += f"\n📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            
            # Отправляем файл
            file = FSInputFile(filename)
            await message.answer_document(
                document=file,
                caption=caption,
                reply_markup=create_main_keyboard()
            )
            
            # Удаляем временный файл
            os.remove(filename)
            logger.info(f"Excel файл {filename} отправлен и удален")
            
        except Exception as e:
            logger.error(f"Ошибка при отправке Excel файла: {e}")
            await message.answer(
                "❌ Произошла ошибка при отправке файла. Попробуйте позже.",
                reply_markup=create_main_keyboard()
            )
    else:
        await message.answer(
            "❌ Произошла ошибка при создании Excel файла. Попробуйте позже.",
            reply_markup=create_main_keyboard()
        )


@dp.message(Command("clear"))
async def cmd_clear(message: Message):
    """Обработчик команды /clear - очистка истории диалога"""
    user_id = message.from_user.id
    clear_history(user_id)
    await message.answer(
        "🗑️ История диалога очищена!\n\n"
        "Начинаем разговор с чистого листа.",
        reply_markup=create_main_keyboard()
    )


@dp.message(Command("history"))
async def cmd_history(message: Message):
    """Обработчик команды /history - показать информацию об истории"""
    user_id = message.from_user.id
    history = get_conversation_history(user_id)
    
    if not history:
        await message.answer(
            "📭 История диалога пуста.",
            reply_markup=create_main_keyboard()
        )
        return
    
    user_msgs = sum(1 for msg in history if msg['role'] == 'user')
    assistant_msgs = sum(1 for msg in history if msg['role'] == 'assistant')
    
    await message.answer(
        f"📊 История диалога:\n\n"
        f"💬 Ваших сообщений: {user_msgs}\n"
        f"🤖 Ответов бота: {assistant_msgs}\n"
        f"📝 Всего в контексте: {len(history)} сообщений",
        reply_markup=create_main_keyboard()
    )


@dp.message(F.voice)
async def handle_voice(message: Message):
    """Обработчик голосовых сообщений"""
    logger.info(f"Получено голосовое сообщение от пользователя {message.from_user.id}")
    
    # Проверяем наличие Groq API ключа
    if not GROQ_API_KEY or GROQ_API_KEY == "your_groq_api_key_here":
        await message.answer(
            "🎤 Голосовые сообщения не настроены.\n\n"
            "📝 Для использования голосовых сообщений необходимо:\n"
            "1. Получить БЕСПЛАТНЫЙ Groq API ключ: https://console.groq.com/keys\n"
            "2. Добавить его в .env файл: GROQ_API_KEY=ваш_ключ\n\n"
            "💬 А пока отправьте ваш вопрос текстом!",
            reply_markup=create_main_keyboard()
        )
        return
    
    if not SYSTEM_PROMPT:
        await message.answer(
            "❌ Извините, бот не настроен правильно. "
            "Системный промпт не загружен. Обратитесь к администратору.",
            reply_markup=create_main_keyboard()
        )
        return
    
    try:
        # Отправляем статус "печатает..."
        await message.answer("🎤 Обрабатываю голосовое сообщение...")
        
        # Скачиваем голосовое сообщение
        file = await bot.get_file(message.voice.file_id)
        file_path = f"voice_{message.voice.file_id}.oga"
        await bot.download_file(file.file_path, file_path)
        
        logger.info(f"Голосовое сообщение скачано: {file_path}")
        
        # Транскрибируем аудио через OpenAI Whisper
        text = await transcribe_audio(file_path)
        
        if not text:
            await message.answer(
                "❌ Не удалось распознать речь. Попробуйте еще раз или отправьте текстовое сообщение.",
                reply_markup=create_main_keyboard()
            )
            return
        
        logger.info(f"Транскрибированный текст: {text}")
        
        # Получаем ответ от LLM с учетом истории
        user_id = message.from_user.id
        response = await get_llm_response(user_id, text)
        
        if response:
            # Отправляем только ответ бота (без расшифровки)
            await send_long_message(message, response)
        else:
            await message.answer(
                "❌ Произошла ошибка при обработке запроса. Попробуйте позже.",
                reply_markup=create_main_keyboard()
            )
            await message.answer(
                "❌ Произошла ошибка при обработке запроса. Попробуйте позже."
            )
            
    except Exception as e:
        logger.error(f"Ошибка при обработке голосового сообщения: {e}")
        await message.answer(
            "❌ Произошла ошибка при обработке голосового сообщения. Попробуйте еще раз.",
            reply_markup=create_main_keyboard()
        )


async def send_long_message(message: Message, text: str) -> None:
    """
    Отправляет длинное сообщение, разбивая его на части если нужно.
    
    Args:
        message: Объект сообщения Telegram
        text: Текст для отправки
    """
    MAX_MESSAGE_LENGTH = 4096
    
    if len(text) <= MAX_MESSAGE_LENGTH:
        await message.answer(text)
        return
    
    # Разбиваем на части по 4000 символов (оставляем запас)
    parts = []
    current_part = ""
    
    for line in text.split('\n'):
        if len(current_part) + len(line) + 1 <= 4000:
            current_part += line + '\n'
        else:
            if current_part:
                parts.append(current_part)
            current_part = line + '\n'
    
    if current_part:
        parts.append(current_part)
    
    # Отправляем части
    for i, part in enumerate(parts, 1):
        prefix = f"📄 Часть {i}/{len(parts)}:\n\n" if len(parts) > 1 else ""
        await message.answer(prefix + part)
        await asyncio.sleep(0.5)  # Небольшая задержка между сообщениями


@dp.message(F.photo)
async def handle_photo(message: Message):
    """Обработчик фотографий"""
    user_id = message.from_user.id
    logger.info(f"Получено фото от пользователя {user_id}")
    
    try:
        # Информируем о получении фото
        await message.answer(
            "📷 Фото получено! Можете добавить текстовый или голосовой комментарий.",
            reply_markup=create_main_keyboard()
        )
        
        # Добавляем информацию о фото в историю
        photo_info = "[Пользователь отправил фото экзаменационной работы/материала]"
        if message.caption:
            photo_info += f"\nПодпись к фото: {message.caption}"
        
        # Добавляем в историю диалога
        add_to_history(user_id, "user", photo_info)
        
        logger.info(f"Фото добавлено в историю пользователя {user_id}")
        
    except Exception as e:
        logger.error(f"Ошибка при обработке фото: {e}")
        await message.answer(
            "❌ Произошла ошибка при обработке фото. Попробуйте еще раз.",
            reply_markup=create_main_keyboard()
        )


@dp.message(F.text)
async def handle_text(message: Message):
    """Обработчик текстовых сообщений"""
    user_id = message.from_user.id
    logger.info(f"Получено текстовое сообщение от пользователя {user_id}")
    
    if not SYSTEM_PROMPT:
        await message.answer(
            "❌ Извините, бот не настроен правильно. "
            "Системный промпт не загружен. Обратитесь к администратору.",
            reply_markup=create_main_keyboard()
        )
        return
    
    try:
        # Отправляем статус "печатает..."
        await message.chat.do("typing")
        
        # Получаем ответ от LLM с учетом истории диалога
        response = await get_llm_response(user_id, message.text)
        
        if response:
            # Отправляем ответ (с автоматическим разбиением на части если нужно)
            await send_long_message(message, response)
        else:
            await message.answer(
                "❌ Произошла ошибка при обработке запроса. Попробуйте позже.",
                reply_markup=create_main_keyboard()
            )
            
    except Exception as e:
        logger.error(f"Ошибка при обработке текстового сообщения: {e}")
        await message.answer(
            "❌ Произошла ошибка при обработке сообщения. Попробуйте еще раз.",
            reply_markup=create_main_keyboard()
        )


async def main():
    """Главная функция запуска бота"""
    # Проверяем наличие необходимых переменных окружения
    if not TELEGRAM_BOT_TOKEN:
        logger.error("TELEGRAM_BOT_TOKEN не установлен в .env файле!")
        sys.exit(1)
    
    if not OPENROUTER_API_KEY:
        logger.error("OPENROUTER_API_KEY не установлен в .env файле!")
        sys.exit(1)
    
    # Загружаем системный промпт
    if not load_system_prompt():
        logger.error("Не удалось загрузить системный промпт. Бот не может работать.")
        sys.exit(1)
    
    # Проверяем настройки голосовых сообщений
    if GROQ_API_KEY and GROQ_API_KEY != "your_groq_api_key_here":
        logger.info("✅ Groq API ключ обнаружен - голосовые сообщения включены (БЕСПЛАТНО!)")
        # Проверяем наличие ffmpeg
        try:
            import subprocess
            subprocess.run(['ffmpeg', '-version'], capture_output=True, check=True)
            logger.info("✅ ffmpeg установлен - обработка аудио доступна")
        except (subprocess.CalledProcessError, FileNotFoundError):
            logger.warning("⚠️  ffmpeg не установлен! Голосовые сообщения не будут работать.")
            logger.warning("   Установите: brew install ffmpeg")
    else:
        logger.warning("⚠️  Groq API ключ не установлен - голосовые сообщения отключены")
        logger.warning("   Получите БЕСПЛАТНЫЙ ключ: https://console.groq.com/keys")
    
    logger.info(f"Запуск бота с моделью: {OPENROUTER_MODEL}")
    logger.info("Бот готов к работе!")
    
    # Запускаем polling
    try:
        await dp.start_polling(bot)
    finally:
        await bot.session.close()


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("Бот остановлен пользователем")